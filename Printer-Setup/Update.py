import os
import configparser
import subprocess
import sys
import tkinter as tk
from tkinter import messagebox

def install(package):
    subprocess.check_call([sys.executable, "-m", "pip", "install", package])

try:
    import openpyxl
except ImportError:
    print("openpyxl is not installed. Installing...")
    install("openpyxl")

def create_School_ini_from_excel(file_path, output_dir):
    # Load the Excel file
    wb = openpyxl.load_workbook(file_path)
    sheet = wb['Assignments']

    # Get the unique values from column A
    unique_values = set([str(row[0]).strip() for row in sheet.iter_rows(min_row=2, values_only=True)])

    # Iterate over the unique values in column A
    for value in unique_values:
        # Skip if value is 'None'
        if value == 'None':
            continue
        # Create a new INI file
        config = configparser.ConfigParser()

        # Iterate over rows starting from the second row
        for row in sheet.iter_rows(min_row=2, values_only=True):

            column_a = str(row[0]).strip()  # Get value from column A

            # Skip rows that don't match the current unique value
            if column_a != value:
                continue
                
            section_title = str(row[1]).strip()  # Get value from column B

            # Create a new section in the INI file
            config[section_title] = {}

            # Set default value of "1" for column F if it is left empty
            secure_print = str(row[5]).strip() or "1"
            config[section_title]['Secure Print'] = secure_print

            # Assign values to Printer 1 and Printer 2
            printer1 = str(row[2]).strip()
            printer2 = str(row[3]).strip()
            if printer1 != 'None':
                config[section_title]['Printer 1'] = printer1
            if printer2 != 'None':
                config[section_title]['Printer 2'] = printer2

            # Split the value in column E by ',' and assign it to Printer 3 and subsequent keys
            printers = str(row[4]).strip().split(',')
            for i, printer in enumerate(printers, start=3):
                key = f'Printer {i}'
                if printer.strip() != 'None':
                    config[section_title][key] = printer.strip()

        # Save the INI file in the specified output directory with the name based on the value in column A
        ini_file_name = os.path.join(output_dir, f'{value}.ini')
        with open(ini_file_name, 'w') as config_file:
            config.write(config_file)

        print(f'Successfully created {ini_file_name}')

def create_config_ini(file_path, output_dir):
    # Load the Excel file
    wb = openpyxl.load_workbook(file_path)
    sheet = wb['Config']

    # Create a new INI file for configuration
    config = configparser.ConfigParser()

    # Iterate over rows starting from the second row
    for row in sheet.iter_rows(min_row=2, values_only=True):
        section_title = str(row[0]).strip()  # Get value from column A
        ip_address = str(row[1]).strip()  # Get value from column B
        location = str(row[2]).strip()  # Get value from column C

        # Create a new section in the INI file
        config[section_title] = {}
        config[section_title]['IP Address'] = ip_address
        config[section_title]['Location'] = location

    # Save the config.ini file in the specified output directory
    config_ini_file = os.path.join(output_dir, 'config.ini')
    with open(config_ini_file, 'w') as config_file:
        config.write(config_file)

    print(f'Successfully created {config_ini_file}')


def create_excel_from_inis(directory_path, output_file):
    # Create a new Excel workbook
    wb = openpyxl.Workbook()

    # Create the "Assignments" tab
    assignment_sheet = wb.active
    assignment_sheet.title = "Assignments"
    assignment_header = ['School Name', 'Room', 'Printer 1', 'Printer 2', 'Other Printers', 'Secure Print']
    assignment_sheet.append(assignment_header)

    # Iterate over the INI files in the directory
    for filename in os.listdir(directory_path):
        if filename.endswith('.ini'):
            ini_file = os.path.join(directory_path, filename)

            # Read the INI file
            config = configparser.ConfigParser()
            config.read(ini_file)

            # Check if the file is 'config.ini'
            if filename == 'config.ini':
                # Create a new tab for 'config.ini'
                config_sheet = wb.create_sheet(title='Config')
                config_header = ['Printer Name', 'IP Address', 'Location']
                config_sheet.append(config_header)

                # Iterate over the sections in the INI file
                for section in config.sections():
                    # Skip the 'config' section
                    if section == 'config':
                        continue

                    # Get the values for each section
                    printer_name = section
                    ip_address = config.get(section, 'IP Address', fallback='')
                    location = config.get(section, 'Location', fallback='')

                    config_row = [printer_name, ip_address, location]
                    config_sheet.append(config_row)
            else:
                # Iterate over the sections in the INI file (excluding 'config')
                for section in config.sections():
                    if section != 'config':
                        # Get the values for each section
                        school_name = filename[:-4]
                        room = section
                        printer1 = config.get(section, 'Printer 1', fallback='')
                        printer2 = config.get(section, 'Printer 2', fallback='')

                        other_printers = ', '.join([
                            config[section].get(f'Printer {i}', '') for i in range(3, 21) if config[section].get(f'Printer {i}') is not None
                        ])

                        secure_print = config.get(section, 'Secure Print', fallback='')

                        # Append the values as a row in the "Assignments" tab
                        assignment_row = [school_name, room, printer1, printer2, other_printers, secure_print]
                        assignment_sheet.append(assignment_row)

    # Remove default "Sheet" tab if it exists
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Save the Excel file
    wb.save(output_file)

    print(f'Successfully created {output_file}')

def create_ini():
    script_directory = os.path.dirname(os.path.abspath(sys.argv[0]))
    # Create Config.ini
    create_config_ini(os.path.join(script_directory, 'Printers.xlsx'), os.path.join(script_directory, 'Files', 'Config'))
    # Create School ini's
    create_School_ini_from_excel(os.path.join(script_directory, 'Printers.xlsx'), os.path.join(script_directory, 'Files', 'Config'))
    messagebox.showinfo("Success", f"INI files created successfully in directory: {os.path.join(script_directory, 'Files', 'Config')}.")

def create_excel():
    script_directory = os.path.dirname(os.path.abspath(sys.argv[0]))
    # Create an excel file from the ini's in the specified directory.
    create_excel_from_inis(os.path.join(script_directory, 'Files', 'Config'), os.path.join(script_directory, 'Printers.xlsx'))
    messagebox.showinfo("Success", f"Excel file created successfully in directory: {os.path.join(script_directory, 'Printers.xlsx')}.")

def main():
    root = tk.Tk()
    root.title("Choose Action")

    explanation_text = tk.Label(root, text="Do you want to create the Printers.xlsx from the INI files or create the INI files from the Printers.xlsx file?")
    explanation_text.pack(pady=10)
    
    def ini_button_clicked():
        create_ini()

    def excel_button_clicked():
        create_excel()

    ini_button = tk.Button(root, text="Create INI", command=ini_button_clicked)
    ini_button.pack(pady=10)

    excel_button = tk.Button(root, text="Create Excel", command=excel_button_clicked)
    excel_button.pack(pady=10)

    root.mainloop()

if __name__ == "__main__":
    main()
